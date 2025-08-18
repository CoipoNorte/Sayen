# p1.py
# Script: Pacientes citados (Rayen APS) — usa funciones compartidas.

from colorama import init
init(autoreset=True)

import os
import time
import re
import datetime
import pandas as pd

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Importa utilidades compartidas
from funciones import (
    iniciar_sesion,
    solicitar_rango_fechas,
    formatear_rut,
    extraer_edad,
    limpiar_nombre,
    MESES_ES,
)

# ==================== Consola ====================

def color_text(text, color="cyan"):
    colors = {
        "cyan": "\033[96m",
        "green": "\033[92m",
        "yellow": "\033[93m",
        "red": "\033[91m",
        "reset": "\033[0m"
    }
    return f"{colors.get(color, '')}{text}{colors['reset']}"

# ==================== Driver silencioso ====================

def crear_driver_silencioso():
    # Reduce ruido de logs de Chrome/Driver en Windows
    os.environ["CHROME_LOG_FILE"] = "NUL"  # en Linux/Mac: '/dev/null'

    options = Options()
    # Mitigaciones Windows + verbosidad mínima
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--log-level=3")
    options.add_argument("--disable-logging")
    options.add_argument("--disable-notifications")
    options.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])
    options.add_experimental_option("useAutomationExtension", False)

    service = Service(log_output=os.devnull)
    return webdriver.Chrome(options=options, service=service)

# ==================== Navegación específica de p1 ====================

def ir_a_pacientes_citados(driver):
    WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.ID, "navbar-main-menu"))).click()
    time.sleep(1)
    WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, "//li[.//span[text()='Box']]"))).click()
    time.sleep(1)
    WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable((By.XPATH, "//a[.//span[contains(text(),'Pacientes citados')]]"))
    ).click()
    time.sleep(3)

def abrir_datepicker(driver):
    WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR,
            ".patient-calendar .react-datepicker__input-container span"))
    ).click()
    WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.CLASS_NAME, "react-datepicker__month"))
    )

def ajustar_mes_anio(driver, anio, mes, meses):
    while True:
        mes_actual = driver.find_element(By.CLASS_NAME,
            "react-datepicker__current-month").text.lower()
        anio_actual = driver.find_element(By.CLASS_NAME,
            "react-datepicker__year-read-view--selected-year").text.strip()

        if mes_actual.startswith(meses[mes-1]) and anio_actual == str(anio):
            break

        if anio_actual != str(anio):
            driver.find_element(By.CLASS_NAME,
                "react-datepicker__year-read-view").click()
            time.sleep(0.5)
            driver.find_element(By.XPATH,
                f"//div[contains(@class, 'react-datepicker__year-option') and text()='{anio}']").click()
            time.sleep(0.5)
        else:
            driver.find_element(By.CLASS_NAME,
                "react-datepicker__month-read-view").click()
            time.sleep(0.5)
            driver.find_element(By.XPATH,
                f"//div[contains(@class, 'react-datepicker__month-option') and text()='{meses[mes-1]}']").click()
            time.sleep(0.5)

def seleccionar_dia(driver, dia, anio, mes):
    xpath_dia = (
        "//div[contains(@class, 'react-datepicker__day')"
        " and not(contains(@class, 'react-datepicker__day--outside-month'))"
        f" and text()='{dia}']"
    )
    try:
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, xpath_dia))
        ).click()
        return True
    except:
        print(color_text(f"Dia {dia}-{mes}-{anio} no disponible. Saltando...", "yellow"))
        return False

def extraer_datos_del_dia(driver, fecha_str, datos):
    filas = driver.find_elements(By.XPATH,
        "//div[contains(@class, 'rt-tr') and @role='row']")

    for idx, fila in enumerate(filas):
        try:
            celdas = fila.find_elements(By.XPATH, ".//div[contains(@class, 'rt-td')]")
            if len(celdas) < 3:
                continue

            nombre = limpiar_nombre(celdas[2].text)
            if not nombre:
                continue

            estado = celdas[1].text.strip().lower()
            tipo_atencion = "NSP" if "no se present" in estado else ""

            fila.click()
            WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.CLASS_NAME, "popover-body"))
            )
            popover = driver.find_element(By.CLASS_NAME, "popover-body").text

            run = ""
            m_run = re.search(r"RUN\s*:? ?([\d\.]+-[\dkK])", popover)
            if m_run:
                run = m_run.group(1).upper()
            else:
                m_run = re.search(r"RUN\s*:? ?(\d+)", popover)
                if m_run:
                    run = formatear_rut(m_run.group(1)).upper()

            sector = ""
            m_sector = re.search(r"sector: (\w+)", popover, re.IGNORECASE)
            if m_sector:
                sector = m_sector.group(1).upper()

            edad_rango = ""
            m_edad = re.search(r"Paciente de: (.+)", popover)
            if m_edad:
                edad_rango = extraer_edad(m_edad.group(1))

            datos.append({
                "FECHA": fecha_str,
                "VACIO1": "",
                "VACIO2": "",
                "SECTOR": sector,
                "NOMBRE": nombre,
                "RUN": run,
                "TIPO DE ATENCIÓN": tipo_atencion,
                "EDAD": edad_rango,
                "DÉFICIT": "",
                "SEXO": "",
            })

        except Exception as e:
            print(color_text(f"Error fila {idx} día {fecha_str}: {e}", "red"))

def guardar_y_colorear(datos, anio, mes, dia_inicio, dia_fin):
    nombre_archivo = f"pacientes_citados_{anio}_{str(mes).zfill(2)}_{dia_inicio:02d}_{dia_fin:02d}.xlsx"
    df = pd.DataFrame(datos, columns=[
        "FECHA", "VACIO1", "VACIO2", "SECTOR", "NOMBRE",
        "RUN", "TIPO DE ATENCIÓN", "EDAD", "DÉFICIT", "SEXO"
    ])
    df.to_excel(nombre_archivo, index=False)
    print(color_text(f"Archivo generado: {nombre_archivo}", "green"))

    wb = load_workbook(nombre_archivo)
    ws = wb.active
    fill_amarillo = PatternFill(start_color="FFF200", end_color="FFF200", fill_type="solid")
    fill_rosa = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

    last_fecha = None
    use_amarillo = True
    for row in range(2, ws.max_row + 1):
        fecha_actual = ws[f"A{row}"].value
        if fecha_actual != last_fecha:
            use_amarillo = not use_amarillo
            last_fecha = fecha_actual
        ws[f"A{row}"].fill = fill_amarillo if use_amarillo else fill_rosa

    wb.save(nombre_archivo)
    print(color_text("Colores aplicados a la columna FECHA.", "green"))

# ==================== Main ====================

def main():
    anio, mes, dia_inicio, dia_fin = solicitar_rango_fechas()

    driver = crear_driver_silencioso()
    iniciar_sesion(driver)
    ir_a_pacientes_citados(driver)

    datos = []
    meses = MESES_ES[:]  # usa los meses compartidos

    for dia in range(dia_inicio, dia_fin + 1):
        try:
            fecha_actual = datetime.date(anio, mes, dia)
        except ValueError:
            continue

        if fecha_actual.weekday() >= 5:
            print(color_text(f"Saltando {fecha_actual} (sábado o domingo)", "yellow"))
            continue

        abrir_datepicker(driver)
        ajustar_mes_anio(driver, anio, mes, meses)

        if not seleccionar_dia(driver, dia, anio, mes):
            continue

        time.sleep(2)

        fecha_str = driver.find_element(By.XPATH, "//strong").text
        extraer_datos_del_dia(driver, fecha_str, datos)

    guardar_y_colorear(datos, anio, mes, dia_inicio, dia_fin)
    driver.quit()

if __name__ == "__main__":
    main()