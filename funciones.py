# funciones.py
# Utilidades compartidas por p1.py y p2.py

"""
Módulo de utilidades compartidas para los scripts p1.py y p2.py.
"""

import os
import re
import time
import unicodedata
import getpass
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service

# Carga .env si existe (opcional)
try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

# ==================== Config / Constantes ====================

BASE_URL = "https://clinico.rayenaps.cl/"

MESES_ES = [
    "enero", "febrero", "marzo", "abril", "mayo", "junio",
    "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"
]

# ==================== Normalización / Texto ====================

def normalize(text: str) -> str:
    text = "" if text is None else str(text)
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = text.upper()
    text = re.sub(r"\s+", " ", text)
    return text.strip()

def norm_simple(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.lower().strip()

def is_empty(val) -> bool:
    if val is None:
        return True
    try:
        import math
        if isinstance(val, float) and math.isnan(val):
            return True
    except Exception:
        pass
    if isinstance(val, str) and val.strip().lower() in {"", "nan", "none", "null"}:
        return True
    return False

# ==================== Utilidades de dominio ====================

def formatear_rut(rut: str) -> str:
    rut = rut.replace(".", "").replace("-", "")
    if len(rut) < 2:
        return rut
    cuerpo = rut[:-1]
    dv = rut[-1].upper()
    cuerpo = cuerpo[::-1]
    partes = [cuerpo[i:i+3] for i in range(0, len(cuerpo), 3)]
    cuerpo_formateado = ".".join(partes)[::-1]
    return f"{cuerpo_formateado}-{dv}"

def extraer_edad(alerta: str) -> str:
    anios = 0
    meses = 0
    m_anios = re.search(r"(\d+)\s*años?", alerta)
    m_meses = re.search(r"(\d+)\s*mes(es)?", alerta)
    if m_anios:
        anios = int(m_anios.group(1))
    if m_meses:
        meses = int(m_meses.group(1))
    total_meses = anios * 12 + meses

    if total_meses == 0:
        m_dias = re.search(r"(\d+)\s*d[ií]as?", alerta)
        if m_dias:
            return "Menor de 7 meses"
        return ""

    if total_meses < 7:
        return "Menor de 7 meses"
    elif 7 <= total_meses <= 11:
        return "7-11 MESES"
    elif 12 <= total_meses <= 17:
        return "12-17 MESES"
    elif 18 <= total_meses <= 23:
        return "18-23 MESES"
    elif 24 <= total_meses <= 47:
        return "24-47 MESES"
    elif 48 <= total_meses <= 59:
        return "48-59 MESES"
    else:
        return "60 MESES O MÁS"

def limpiar_nombre(valor) -> str:
    s = "" if valor is None else str(valor)
    s = re.sub(r"[（(][^）)]*[）)]", "", s, flags=re.UNICODE)
    s = re.sub(r"\s+", " ", s, flags=re.UNICODE).strip()
    return s

def edad_menor_4_meses(edad_str: str) -> bool:
    anios = 0
    meses = 0
    m_anios = re.search(r"(\d+)\s*años?", edad_str)
    m_meses = re.search(r"(\d+)\s*mes(es)?", edad_str)
    if m_anios:
        anios = int(m_anios.group(1))
    if m_meses:
        meses = int(m_meses.group(1))
    total_meses = anios * 12 + meses
    return total_meses < 4

# ==================== Mapeos de atención / análisis ====================

TIPOS_MAP = {
    "ATENCION PRESENCIAL":           "ASISTE",
    "NO SE PRESENT":                 "NSP",
    "NSP":                           "NSP",
    "LLAMADO TELEFONICO EFECTIVO":   "LLAMADO TELEFÓNICO EFECTIVO",
    "LLAMADO TELEFONICO INEFECTIVO": "LLAMADO TELEFÓNICO INEFECTIVO",
    "VIDEOLLAMADA EFECTIVA":         "VIDEOLLAMADA EFECTIVA",
    "VIDEOLLAMADA INEFECTIVA":       "VIDEOLLAMADA INEFECTIVA",
    "MENSAJERIA DE TEXTO":           "MENSAJERÍA DE TEXTO",
    "VISITA DOMICILARIA INTEGRAL EFECTIVA":   "V.D.I EFECTIVA",
    "VISITA DOMICILARIA INTEGRAL INEFECTIVA": "V.D.I INEFECTIVA",
    "EGRESO ADMINISTRATIVO":         "EGRESO ADMINISTRATIVO",
}

DEFICITS = ["REZAGO", "RIESGO", "RETRASO", "RBPS", "NANEAS"]

def detect_tipo_y_deficit(raw_text: str):
    txt = normalize(raw_text)
    tipo = ""
    for clave in sorted(TIPOS_MAP, key=len, reverse=True):
        if clave in txt:
            tipo = TIPOS_MAP[clave]
            break
    deficit = next((d for d in DEFICITS if d in txt), "")
    return tipo, deficit

def normalizar_tipo_excel(valor):
    if is_empty(valor):
        return ""
    txt = normalize(valor)
    for k, v in TIPOS_MAP.items():
        if k in txt:
            return v
    return str(valor).strip()

# ==================== Credenciales y login ====================

def get_credentials():
    location = os.getenv("RAYEN_LOCATION")
    username = os.getenv("RAYEN_USERNAME")
    password = os.getenv("RAYEN_PASSWORD")

    if not location:
        location = input("Ubicación (RAYEN_LOCATION): ").strip()
    if not username:
        username = input("Usuario (RAYEN_USERNAME): ").strip()
    if not password:
        password = getpass.getpass("Contraseña (RAYEN_PASSWORD): ").strip()

    return location, username, password

def iniciar_sesion(driver):
    location, username, password = get_credentials()

    driver.get(BASE_URL)
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, "location")))
    driver.find_element(By.ID, "location").send_keys(location)
    driver.find_element(By.ID, "username").send_keys(username)
    driver.find_element(By.ID, "password").send_keys(password)
    WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Ingresar') and not(@disabled)]"))
    ).click()
    time.sleep(5)

def esperar_loader_y_modal(driver, wait):
    try:
        wait.until(EC.invisibility_of_element_located((By.CLASS_NAME, "cache-loading")))
    except Exception:
        pass
    try:
        modal = driver.find_element(By.XPATH, "//div[contains(@class,'modal') and contains(@class,'show')]")
        for btn in modal.find_elements(By.XPATH, ".//button"):
            if btn.is_displayed():
                btn.click()
                time.sleep(1)
    except Exception:
        pass

# ==================== Entrada común ====================

def solicitar_rango_fechas():
    anio = int(input("Año (ej: 2025): ").strip())
    mes = int(input("Mes (número, ej: 7 para julio): ").strip())
    dia_inicio = int(input("Día de inicio: ").strip())
    dia_fin = int(input("Día de fin: ").strip())
    return anio, mes, dia_inicio, dia_fin

# ==================== Driver compartido (silencioso) ====================

def crear_driver_silencioso(headless: bool | None = None) -> webdriver.Chrome:
    os.environ["CHROME_LOG_FILE"] = "NUL" if os.name == "nt" else "/dev/null"

    if headless is None:
        headless = str(os.getenv("HEADLESS", "false")).lower() in {"1", "true", "yes"}

    options = Options()
    if headless:
        options.add_argument("--headless=new")
        options.add_argument("--window-size=1920,1080")

    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--log-level=3")
    options.add_argument("--disable-logging")
    options.add_argument("--disable-notifications")
    options.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])
    options.add_experimental_option("useAutomationExtension", False)

    service = Service(log_output=os.devnull)
    return webdriver.Chrome(options=options, service=service)

# ==================== Escritura en Excel (in place) ====================

def actualizar_excel_inplace(
    excel_path: str,
    df: pd.DataFrame,
    columnas_objetivo: Dict[str, List[str]],
    only_if_empty: bool = True
) -> Dict[str, int]:
    """
    Actualiza el Excel 'excel_path' escribiendo SOLO celdas vacías para las columnas objetivo.
    Si una columna no existe en el Excel, la crea al final con el encabezado canónico.

    Args:
      excel_path (str): Ruta del Excel a actualizar.
      df (pd.DataFrame): DataFrame con las columnas a escribir (mismos renglones que el Excel).
      columnas_objetivo (dict): {'CANONICO': ['alias1','alias2',...], ...}
      only_if_empty (bool): Si True, solo escribe cuando la celda está vacía (None/"").

    Returns:
      dict[str,int]: Cantidad de celdas escritas por columna canónica.
    """
    wb = load_workbook(excel_path)
    ws = wb.active

    # Mapear encabezados del Excel (fila 1) -> índice de columna
    header_map = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        key = str(val).strip().upper() if val is not None else ""
        if key:
            header_map[key] = col

    escritos_por_col = {}

    # Asegurar que todas las columnas canónicas existan (o crearlas)
    for canonico, aliases in columnas_objetivo.items():
        col_idx = None
        # Buscar por alias
        for alias in aliases + [canonico]:
            alias_up = alias.strip().upper()
            if alias_up in header_map:
                col_idx = header_map[alias_up]
                break
        # Si no existe, crear nueva al final
        if col_idx is None:
            col_idx = ws.max_column + 1
            ws.cell(row=1, column=col_idx, value=canonico)
            header_map[canonico.upper()] = col_idx

        escritos = 0
        # Escribir valores fila a fila (DataFrame debe alinear con el Excel: fila 2 -> df.iloc[0])
        for i in range(2, ws.max_row + 1):
            df_idx = i - 2
            if df_idx >= len(df):
                break
            nuevo_valor = df.at[df.index[df_idx], canonico] if canonico in df.columns else None
            # Si DataFrame tiene el valor y no está vacío
            if not is_empty(nuevo_valor):
                celda = ws.cell(row=i, column=col_idx)
                if (not only_if_empty) or (celda.value in (None, "")):
                    celda.value = nuevo_valor
                    escritos += 1
        escritos_por_col[canonico] = escritos

    wb.save(excel_path)
    return escritos_por_col