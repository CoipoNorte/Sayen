"""
Servicio para manejo de archivos Excel.
"""
from pathlib import Path
from typing import List, Dict, Optional
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from src.domain.models import Paciente
from src.core.logging import get_logger
from src.core.exceptions import ExcelProcessingError
from src.core.utils import is_empty


logger = get_logger(__name__)


class ExcelService:
    """Servicio para operaciones con archivos Excel."""
    
    @staticmethod
    def save_patients(
        patients: List[Paciente], 
        filename: str,
        apply_colors: bool = True
    ) -> Path:
        """
        Guarda una lista de pacientes en un archivo Excel.
        
        Args:
            patients: Lista de pacientes
            filename: Nombre del archivo
            apply_colors: Si aplicar colores alternados por fecha
            
        Returns:
            Ruta del archivo guardado
        """
        try:
            # Convierte pacientes a DataFrame
            data = [p.to_dict() for p in patients]
            df = pd.DataFrame(data)
            
            # Agrega columnas vacías si es necesario
            for col in ["VACIO1", "VACIO2"]:
                if col not in df.columns:
                    df.insert(1, col, "")
            
            # Ordena columnas
            column_order = [
                "FECHA", "VACIO1", "VACIO2", "SECTOR", "NOMBRE",
                "RUN", "TIPO DE ATENCIÓN", "EDAD", "DÉFICIT", "SEXO", "CONSEJERIA"
            ]
            df = df.reindex(columns=[c for c in column_order if c in df.columns])
            
            # Guarda Excel
            filepath = Path(filename)
            df.to_excel(filepath, index=False)
            logger.info(f"Excel guardado: {filepath}")
            
            # Aplica colores si se solicita
            if apply_colors:
                ExcelService._apply_date_colors(filepath)
            
            return filepath
            
        except Exception as e:
            logger.error(f"Error guardando Excel: {e}")
            raise ExcelProcessingError(f"Error guardando Excel: {e}")
    
    @staticmethod
    def _apply_date_colors(filepath: Path) -> None:
        """Aplica colores alternados por fecha en la columna FECHA."""
        try:
            wb = load_workbook(filepath)
            ws = wb.active
            
            # Define colores
            yellow_fill = PatternFill(start_color="FFF200", end_color="FFF200", fill_type="solid")
            pink_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            
            # Alterna colores por fecha
            last_date = None
            use_yellow = True
            
            for row in range(2, ws.max_row + 1):
                current_date = ws[f"A{row}"].value
                
                if current_date != last_date:
                    use_yellow = not use_yellow
                    last_date = current_date
                
                ws[f"A{row}"].fill = yellow_fill if use_yellow else pink_fill
            
            wb.save(filepath)
            logger.debug("Colores aplicados al Excel")
            
        except Exception as e:
            logger.warning(f"No se pudieron aplicar colores: {e}")
    
    @staticmethod
    def load_patients(filepath: str) -> pd.DataFrame:
        """
        Carga pacientes desde un archivo Excel.
        
        Args:
            filepath: Ruta del archivo Excel
            
        Returns:
            DataFrame con los datos
        """
        try:
            df = pd.read_excel(filepath)
            df.columns = df.columns.str.strip().str.upper()
            logger.info(f"Excel cargado: {filepath}")
            return df
            
        except Exception as e:
            logger.error(f"Error cargando Excel: {e}")
            raise ExcelProcessingError(f"Error cargando Excel: {e}")
    
    @staticmethod
    def update_excel_inplace(
        filepath: str,
        updates: pd.DataFrame,
        column_mapping: Dict[str, List[str]],
        only_empty: bool = True
    ) -> Dict[str, int]:
        """
        Actualiza un archivo Excel in-place.
        
        Args:
            filepath: Ruta del archivo Excel
            updates: DataFrame con actualizaciones
            column_mapping: Mapeo de columnas canónicas a alias
            only_empty: Solo actualizar celdas vacías
            
        Returns:
            Diccionario con cantidad de celdas actualizadas por columna
        """
        try:
            wb = load_workbook(filepath)
            ws = wb.active
            
            # Mapea encabezados existentes
            header_map = {}
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=1, column=col).value
                if val:
                    header_map[str(val).strip().upper()] = col
            
            stats = {}
            
            for canonical, aliases in column_mapping.items():
                # Busca columna existente o crea nueva
                col_idx = None
                for alias in aliases + [canonical]:
                    if alias.upper() in header_map:
                        col_idx = header_map[alias.upper()]
                        break
                
                if col_idx is None:
                    col_idx = ws.max_column + 1
                    ws.cell(row=1, column=col_idx, value=canonical)
                    header_map[canonical.upper()] = col_idx
                
                # Actualiza valores
                updated = 0
                for i in range(2, min(ws.max_row + 1, len(updates) + 2)):
                    df_idx = i - 2
                    if df_idx >= len(updates):
                        break
                    
                    if canonical in updates.columns:
                        new_value = updates.at[updates.index[df_idx], canonical]
                        
                        if not is_empty(new_value):
                            cell = ws.cell(row=i, column=col_idx)
                            
                            if not only_empty or is_empty(cell.value):
                                cell.value = new_value
                                updated += 1
                
                stats[canonical] = updated
            
            wb.save(filepath)
            logger.info(f"Excel actualizado: {filepath}")
            return stats
            
        except Exception as e:
            logger.error(f"Error actualizando Excel: {e}")
            raise ExcelProcessingError(f"Error actualizando Excel: {e}")